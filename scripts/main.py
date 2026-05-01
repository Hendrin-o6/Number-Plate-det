import cv2
import pytesseract
import numpy as np
import re
from ultralytics import YOLO
import datetime
import os
try:
    from openpyxl import Workbook, load_workbook
    _EXCEL_AVAILABLE = True
except Exception:
    print('openpyxl not installed; Excel logging will be disabled.')
    _EXCEL_AVAILABLE = False

# ✅ Tesseract path (Windows)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ✅ Load YOLOv8 model (with error tracing)
try:
    model = YOLO(r"C:\Users\hendr\OneDrive\Desktop\Number Plate\weights\numberplate.pt")
except Exception as e:
    import traceback
    print("Failed to load YOLO model:\n", e)
    traceback.print_exc()
    raise SystemExit(1)

# ✅ Logging setup
excel_path = r"C:\Users\hendr\OneDrive\Desktop\Number Plate\log\log.xlsx"
plate_folder = r"C:\Users\hendr\OneDrive\Desktop\Number Plate\log\plates"
os.makedirs(plate_folder, exist_ok=True)

# ✅ Create Excel file if it doesn't exist and openpyxl is available
if _EXCEL_AVAILABLE:
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detections"
        ws.append(["Timestamp", "Plate Text", "Image Path"])
        wb.save(excel_path)

# ✅ Open laptop webcam
cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("Webcam not accessible.")
    exit()

while True:
    ret, frame = cap.read()
    if not ret:
        print("Frame capture failed.")
        break

    # Run inference (support different ultralytics API styles)
    try:
        results = model(frame)
    except TypeError:
        # some ultralytics versions expect predict()
        results = model.predict(source=frame)
    except Exception as e:
        print("Model inference failed:", e)
        import traceback
        traceback.print_exc()
        continue

    for r in results:
        for box in r.boxes:
            # confidence may be a tensor/array; coerce to float safely
            conf = 0.0
            try:
                # Try indexing into conf (tensor-like)
                conf = float(box.conf[0])
            except (IndexError, TypeError):
                try:
                    # Try direct float conversion
                    conf = float(box.conf)
                except (TypeError, ValueError):
                    conf = 0.0

            if conf < 0.5:
                continue

            # xyxy may be tensor-like; convert to ints safely
            x1, y1, x2, y2 = None, None, None, None
            try:
                # Try indexing xyxy[0]
                coords = box.xyxy[0]
                x1, y1, x2, y2 = [int(float(x)) for x in coords]
            except (IndexError, TypeError, ValueError):
                try:
                    # Fallback: try direct iteration or slicing
                    coords = list(box.xyxy) if hasattr(box.xyxy, '__iter__') else [box.xyxy]
                    if len(coords) >= 4:
                        x1, y1, x2, y2 = [int(float(x)) for x in coords[:4]]
                except (TypeError, ValueError, IndexError):
                    print("Couldn't parse box coordinates, skipping box")
                    continue

            if x1 is None or y1 is None or x2 is None or y2 is None:
                print("Invalid coordinates, skipping box")
                continue

            # crop and check size
            y1_safe = max(0, int(y1))
            y2_safe = max(0, int(y2))
            x1_safe = max(0, int(x1))
            x2_safe = max(0, int(x2))
            
            if y2_safe <= y1_safe or x2_safe <= x1_safe:
                print(f"Invalid crop bounds: y1={y1_safe} y2={y2_safe} x1={x1_safe} x2={x2_safe}, skipping")
                continue
            
            plate_img = frame[y1_safe:y2_safe, x1_safe:x2_safe]
            if plate_img.size == 0:
                print("Empty crop, skipping")
                continue

            # ✅ OCR preprocessing
            gray = cv2.cvtColor(plate_img, cv2.COLOR_BGR2GRAY)
            enhanced = cv2.bilateralFilter(gray, 11, 17, 17)
            thresh = cv2.adaptiveThreshold(enhanced, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                           cv2.THRESH_BINARY, 11, 2)

            # ✅ OCR with hardened error handling
            def perform_ocr(img):
                """Try several preprocessing steps and tesseract configs, return best text or empty string."""
                candidates = []
                try:
                    configs = [
                        '--oem 3 --psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789',
                        '--oem 3 --psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789',
                        '--oem 3 --psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
                    ]
                    scales = [1.0, 1.5, 2.0]

                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    blur = cv2.bilateralFilter(gray, 9, 75, 75)

                    for s in scales:
                        if s != 1.0:
                            h, w = blur.shape[:2]
                            scaled = cv2.resize(blur, (int(w * s), int(h * s)), interpolation=cv2.INTER_CUBIC)
                        else:
                            scaled = blur

                        th_otsu = cv2.threshold(scaled, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
                        th_adapt = cv2.adaptiveThreshold(scaled, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                                         cv2.THRESH_BINARY, 11, 2)

                        for prep_img in (scaled, th_otsu, th_adapt):
                            for cfg in configs:
                                try:
                                    data = pytesseract.image_to_data(prep_img, output_type=pytesseract.Output.DICT, config=cfg)
                                except pytesseract.TesseractNotFoundError:
                                    return "OCR_FAILED"
                                except Exception as e:
                                    print(f"  Tesseract error: {e}")
                                    continue

                                try:
                                    texts = [t.strip() for t in data.get('text', []) if t.strip()]
                                    conf_list = data.get('conf', [])
                                    confs = []
                                    for c in conf_list:
                                        try:
                                            val = int(str(c).strip())
                                            if val >= 0:
                                                confs.append(val)
                                        except (ValueError, AttributeError, TypeError):
                                            pass
                                    
                                    text_join = " ".join(texts)
                                    mean_conf = (sum(confs) / len(confs)) if confs else 0
                                    candidates.append((text_join, mean_conf))
                                except Exception as e:
                                    print(f"  Error parsing OCR data: {e}")
                                    continue

                    candidates = [c for c in candidates if c[0]]
                    if not candidates:
                        return ""
                    best = max(candidates, key=lambda x: x[1])
                    cleaned = re.sub(r'[^A-Z0-9\- ]', '', best[0].upper())
                    return cleaned.strip()
                except pytesseract.TesseractNotFoundError:
                    return "OCR_FAILED"
                except Exception as e:
                    print(f"  perform_ocr error: {e}")
                    import traceback
                    traceback.print_exc()
                    return ""

            text = perform_ocr(plate_img)
            print("OCR Text:", repr(text))

            # Always save crop so we can inspect failures; include "_EMPTY" in name when OCR returns empty
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            filename = timestamp.replace(":", "-").replace(" ", "_") + ".jpg"
            image_path = os.path.join(plate_folder, filename)
            try:
                cv2.imwrite(image_path, plate_img)
            except Exception as e:
                print("Failed to write image:", e)

            # ✅ Save image and log to Excel only if text is found and openpyxl is available
            if text and _EXCEL_AVAILABLE:
                try:
                    wb = load_workbook(excel_path)
                    ws = wb["Detections"]
                    ws.append([timestamp, text, image_path])
                    wb.save(excel_path)
                    print(f"Logged: {timestamp}, {text}")
                except Exception as e:
                    print("Failed to write to Excel:", e)
                    import traceback
                    traceback.print_exc()
            else:
                print("No readable text found. Saved crop for inspection.")

            # Also append a simple text log for quick debugging
            try:
                txtlog = os.path.join(os.path.dirname(excel_path), 'log.txt')
                with open(txtlog, 'a', encoding='utf-8') as f:
                    f.write(f"{timestamp}\t{text}\t{image_path}\n")
            except Exception as e:
                print("Failed to append to log.txt:", e)

            # ✅ Annotate frame
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(frame, text if text else "No Text", (x1, y1 - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

    cv2.imshow("Laptop Cam Detection", frame)
    if cv2.waitKey(1) == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
